"""Export filtered jobs to Excel, with read-back for deduplication."""

import os
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

COLUMNS = ["company_name", "job_title", "location", "url", "date_posted", "score", "source"]
COL_WIDTHS = [28, 45, 30, 65, 14, 8, 16]

# openpyxl requires 8-char ARGB — prefix FF for fully opaque
HEADER_BG = "FF2E4057"
HEADER_FG = "FFFFFFFF"

# Cycling pastel backgrounds for each date group (FF = fully opaque)
PASTEL_COLORS = ["FFC5D8F5", "FFC8EAD5", "FFFFF0B3", "FFFADDC9", "FFDDD5F5", "FFC5EDE8"]


def load_existing_jobs(path: str) -> list[dict]:
    """
    Read jobs already saved in an Excel file.
    Returns an empty list if the file doesn't exist or can't be read.
    Skips date-group header rows (where url and job_title are both empty).
    """
    if not Path(path).exists():
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return []
    if not rows:
        return []
    # Normalise headers: lowercase + spaces → underscores so keys match COLUMNS
    # e.g. "Company Name" → "company_name", "Date Posted" → "date_posted"
    header = [
        str(c).lower().replace(" ", "_") if c else ""
        for c in rows[0]
    ]
    jobs = []
    for row in rows[1:]:
        job = {header[i]: (row[i] or "") for i in range(min(len(header), len(row)))}
        # Skip date-group header rows (merged label rows have no url or job_title)
        if not job.get("url") and not job.get("job_title"):
            continue
        jobs.append(job)
    return jobs


def export_to_excel(jobs: list[dict], path: str) -> None:
    jobs = sorted(jobs, key=lambda j: j.get("date_posted") or "", reverse=True)
    parent = os.path.dirname(path) or "."
    os.makedirs(parent, exist_ok=True)

    # Group by date_posted
    date_groups: dict[str, list[dict]] = defaultdict(list)
    for job in jobs:
        date_groups[job.get("date_posted") or "Unknown"].append(job)
    # Sort real dates newest-first; push "Unknown" to the bottom
    sorted_dates = sorted(
        date_groups.keys(),
        key=lambda d: (d == "Unknown", d),
        reverse=True,
    )
    sorted_dates = [d for d in sorted_dates if d != "Unknown"] + (
        ["Unknown"] if "Unknown" in date_groups else []
    )

    wb = Workbook()
    ws = wb.active

    # Column header row
    header_fill = PatternFill("solid", start_color=HEADER_BG)
    header_font = Font(bold=True, color=HEADER_FG, name="Arial", size=10)
    for c, col in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=col.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    current_row = 2
    data_font = Font(name="Arial", size=9)

    for color_idx, date in enumerate(sorted_dates):
        group_jobs = date_groups[date]
        color_hex = PASTEL_COLORS[color_idx % len(PASTEL_COLORS)]
        date_fill = PatternFill("solid", start_color=color_hex)

        # Date group header row
        n = len(group_jobs)
        label = f"  {date}    {n} job{'s' if n != 1 else ''}"
        date_cell = ws.cell(row=current_row, column=1, value=label)
        date_cell.font = Font(bold=True, name="Arial", size=10)
        date_cell.fill = date_fill
        date_cell.alignment = Alignment(vertical="center")
        for c in range(2, len(COLUMNS) + 1):
            ws.cell(row=current_row, column=c).fill = date_fill
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(COLUMNS),
        )
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # Job rows
        url_col = COLUMNS.index("url") + 1  # 1-based column index for URL
        for job in group_jobs:
            for c, col in enumerate(COLUMNS, start=1):
                value = job.get(col, "")
                cell = ws.cell(row=current_row, column=c, value=value)
                if c == url_col and value:
                    cell.hyperlink = value
                    cell.font = Font(name="Arial", size=9,
                                     color="FF0563C1", underline="single")
                else:
                    cell.font = data_font
                cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[current_row].height = 15
            current_row += 1

    # Fixed column widths
    for c, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = "A2"
    wb.save(path)
