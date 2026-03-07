#!/usr/bin/env python3
"""
Mark a job as applied so it's excluded from future email digests.

Usage:
    python mark_applied.py <job_url>
    python mark_applied.py --list          # show all applied jobs
    python mark_applied.py --remove <url>  # un-mark a job

Examples:
    python mark_applied.py "https://boards.greenhouse.io/stripe/jobs/12345"
    python mark_applied.py --list
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

APPLIED_PATH = Path("data/applied_jobs.json")


def load() -> list[dict]:
    if not APPLIED_PATH.exists():
        return []
    try:
        return json.loads(APPLIED_PATH.read_text())
    except (json.JSONDecodeError, OSError):
        return []


def save(jobs: list[dict]) -> None:
    APPLIED_PATH.parent.mkdir(parents=True, exist_ok=True)
    APPLIED_PATH.write_text(json.dumps(jobs, indent=2))


def _find_job_in_excel(url: str) -> dict | None:
    """Look up job metadata from the output Excel by URL."""
    excel = Path("output/entry_roles.xlsx")
    if not excel.exists():
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(excel, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return None
        header = [str(c).lower().replace(" ", "_") if c else "" for c in rows[0]]
        for row in rows[1:]:
            job = {header[i]: (row[i] or "") for i in range(min(len(header), len(row)))}
            if job.get("url") == url:
                return job
    except Exception:
        pass
    return None


def mark_applied(url: str) -> None:
    applied = load()
    existing_urls = {j["job_url"] for j in applied}

    if url in existing_urls:
        print(f"Already marked as applied: {url}")
        return

    # Try to enrich with metadata from the Excel
    meta = _find_job_in_excel(url) or {}
    entry = {
        "company":      meta.get("company_name") or "Unknown",
        "role":         meta.get("job_title")    or "Unknown",
        "job_url":      url,
        "date_applied": datetime.now().strftime("%Y-%m-%d"),
        "source":       meta.get("source")       or "unknown",
    }
    applied.append(entry)
    save(applied)
    print(f"✓ Marked as applied: {entry['role']} @ {entry['company']}")
    print(f"  URL: {url}")
    print(f"  Date: {entry['date_applied']}")


def list_applied() -> None:
    applied = load()
    if not applied:
        print("No jobs marked as applied yet.")
        return
    print(f"\n{'#':<4} {'Company':<30} {'Role':<40} {'Date':<12} Source")
    print("-" * 100)
    for i, j in enumerate(applied, 1):
        print(f"{i:<4} {j.get('company',''):<30} {j.get('role',''):<40} "
              f"{j.get('date_applied',''):<12} {j.get('source','')}")
    print(f"\nTotal: {len(applied)} applied")


def remove_applied(url: str) -> None:
    applied = load()
    before = len(applied)
    applied = [j for j in applied if j.get("job_url") != url]
    if len(applied) == before:
        print(f"URL not found in applied list: {url}")
        return
    save(applied)
    print(f"✓ Removed from applied list: {url}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Track jobs you've applied to",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("url", nargs="?", help="Job URL to mark as applied")
    parser.add_argument("--list", "-l", action="store_true", help="List all applied jobs")
    parser.add_argument("--remove", "-r", metavar="URL", help="Remove a URL from applied list")
    args = parser.parse_args()

    if args.list:
        list_applied()
    elif args.remove:
        remove_applied(args.remove)
    elif args.url:
        mark_applied(args.url)
    else:
        parser.print_help()
        sys.exit(1)


if __name__ == "__main__":
    main()
